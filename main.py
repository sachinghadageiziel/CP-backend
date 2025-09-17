import os
import re
import uuid
import pandas as pd
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pdf2docx import Converter
from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create FastAPI app
app = FastAPI()

#  Enable CORS for React frontend
# app.add_middleware(
#     CORSMiddleware,
#     allow_origins=[
#         "http://localhost:3000",
#         "http://127.0.0.1:3000",
#         "https://gilded-gaufre-2eadbb.netlify.app/",
#         "https://cp-backend-6n14.onrender.com"
#     ],  
    
#     allow_credentials=True,
#     allow_methods=["*"],
#     allow_headers=["*"],
# )

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # allow everything
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Storage folder
OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# STEP 1: Convert PDF -> Word
def pdf_to_word(pdf_file, word_file):
    cv = Converter(pdf_file)
    cv.convert(word_file, start=0, end=None)
    cv.close()

# STEP 2: Extract Requirements from Word
def extract_requirements_from_docx(docx_file):
    doc = Document(docx_file)
    requirements = []
    current_section = ""

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        sec_match = re.match(r"^(\d+(\.\d+)+)", text)
        if sec_match:
            current_section = sec_match.group(1)

        if re.search(r"\b(shall)\b", text, re.I):
            keyword = re.search(r"\b(shall|should|must)\b", text, re.I).group(1).lower()
            requirements.append({
                "Requirement ID": f"REQ-{len(requirements)+1}",
                "Section": current_section,
                "Keyword": keyword,
                "Requirement Text": text
            })
    return requirements

# STEP 3: Save to Excel with Formatting
def save_to_excel(requirements, output_file):
    df = pd.DataFrame(requirements)
    df = df[["Requirement ID", "Section", "Keyword", "Requirement Text"]]
    df.to_excel(output_file, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    fill_alt = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = fill_alt

    wb.save(output_file)

# ---------- API Routes ----------
@app.get("/")
def root():
    return {"message": "Backend is running 🚀"}

@app.post("/extract")
async def extract_requirements(file: UploadFile = File(...)):
    pdf_path = os.path.join(OUTPUT_DIR, f"{uuid.uuid4()}.pdf")
    with open(pdf_path, "wb") as f:
        f.write(await file.read())

    word_path = pdf_path.replace(".pdf", ".docx")
    pdf_to_word(pdf_path, word_path)

    reqs = extract_requirements_from_docx(word_path)
    if not reqs:
        return JSONResponse({"error": "No requirements found."}, status_code=400)

    excel_path = pdf_path.replace(".pdf", ".xlsx")
    save_to_excel(reqs, excel_path)

    return {
        "count": len(reqs),
        "download_url": f"https://cp-backend-6n14.onrender.com/download/{os.path.basename(excel_path)}"
    }

@app.get("/download/{filename}")
async def download_file(filename: str):
    file_path = os.path.join(OUTPUT_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, filename=filename)
    return JSONResponse({"error": "File not found"}, status_code=404)
